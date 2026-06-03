from __future__ import annotations

import csv
import hashlib
import json
import re
from datetime import date, datetime, timedelta
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ubs_monitor.db import is_postgres_connection


DEFAULT_WORKBOOK_PATH = Path("/Users/felipeoliveira/Downloads/ Gestantes e Puerperas Guajiru.xlsx")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def slug_text(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", clean_text(value).lower()).strip("_")


def digits_only(value: Any) -> str:
    return re.sub(r"\D+", "", clean_text(value))


def parse_excel_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_text(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def parse_boolish(value: Any) -> int:
    text = clean_text(value).lower()
    if not text:
        return 0
    return 1 if text in {"1", "sim", "yes", "true", "realizado", "realizada", "em dia"} else 0


def normalize_risk(value: Any) -> str:
    text = clean_text(value).lower()
    if "alto" in text:
        return "Alto risco"
    if "inter" in text or "moderad" in text:
        return "Risco intermediario"
    if "baixo" in text or "habitual" in text:
        return "Baixo risco"
    return "Sem classificacao"


def extract_weeks(value: Any) -> int | None:
    text = clean_text(value).lower()
    match = re.search(r"(\d{1,2})", text)
    if not match:
        return None
    weeks = int(match.group(1))
    return weeks if 1 <= weeks <= 45 else None


def estimate_dates(last_consultation_iso: str | None, gestational_weeks: int | None) -> tuple[str | None, str | None]:
    if not last_consultation_iso or not gestational_weeks:
        return None, None
    consultation_date = date.fromisoformat(last_consultation_iso)
    dum = consultation_date - timedelta(weeks=gestational_weeks)
    dpp = dum + timedelta(weeks=40)
    return dum.isoformat(), dpp.isoformat()


def infer_status(record: dict) -> str:
    if record.get("actual_birth_date"):
        return "puerpera"
    text = " ".join(
        clean_text(record.get(key))
        for key in ("status", "desfecho", "gestational_age_label")
        if record.get(key) is not None
    ).lower()
    if any(token in text for token in ("puerper", "parto realizado", "alta")):
        return "puerpera"
    if "encerr" in text:
        return "encerrado"
    return "gestante"


def composite_header(top_value: Any, bottom_value: Any, last_top: str) -> tuple[str, str]:
    top = clean_text(top_value) or last_top
    bottom = clean_text(bottom_value)
    if top and bottom:
        return f"{top} {bottom}", top
    return top or bottom, top


def make_row_hash(payload: dict) -> str:
    serialized = json.dumps(payload, sort_keys=True, ensure_ascii=False, default=str)
    return hashlib.sha256(serialized.encode("utf-8")).hexdigest()


def make_source_key(patient: dict) -> str:
    cpf = digits_only(patient.get("cpf"))
    if cpf:
        return f"cpf:{cpf}"
    cns = digits_only(patient.get("cns"))
    if cns:
        return f"cns:{cns}"
    external_code = clean_text(patient.get("external_code"))
    if external_code:
        return f"external:{external_code.lower()}"
    birth_date = patient.get("birth_date") or ""
    mother_name = slug_text(patient.get("mother_name"))
    name = slug_text(patient.get("name"))
    return f"composite:{name}|{birth_date}|{mother_name}"


def build_event_key(source_system: str, source_key: str, event_type: str, event_date: str) -> str:
    payload = f"{source_system}|{source_key}|{event_type}|{event_date}"
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def add_event(events: list[dict], *, event_type: str, event_date: str | None, professional: str | None, notes: str | None, metadata: dict) -> None:
    if not event_date:
        return
    events.append(
        {
            "event_type": event_type,
            "event_date": event_date,
            "professional": professional,
            "notes": notes,
            "metadata": metadata,
        }
    )


def parse_registro_sheet(workbook_bytes: bytes) -> list[dict]:
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    sheet = workbook["Registro pac."] if "Registro pac." in workbook.sheetnames else workbook[workbook.sheetnames[0]]

    records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = clean_text(row[1] if len(row) > 1 else None)
        if not name:
            continue

        external_code = clean_text(row[0] if len(row) > 0 else None) or None
        locality = clean_text(row[2] if len(row) > 2 else None) or None
        risk_level = normalize_risk(row[3] if len(row) > 3 else None)
        last_consultation_date = parse_excel_date(row[4] if len(row) > 4 else None)
        gestational_age_label = clean_text(row[5] if len(row) > 5 else None) or None
        gestational_weeks = extract_weeks(gestational_age_label)
        notes = clean_text(row[6] if len(row) > 6 else None) or None
        professional = clean_text(row[7] if len(row) > 7 else None) or None
        active_search = parse_boolish(row[8] if len(row) > 8 else None)
        dum, dpp = estimate_dates(last_consultation_date, gestational_weeks)

        patient = {
            "external_code": external_code,
            "name": name,
            "locality": locality,
            "risk_level": risk_level,
            "status": "gestante",
            "gestational_weeks": gestational_weeks,
            "gestational_age_label": gestational_age_label,
            "dum": dum,
            "dpp": dpp,
            "last_consultation_date": last_consultation_date,
            "last_professional": professional,
            "active_search": active_search,
            "notes": notes,
            "source": "registro_pac",
        }

        events: list[dict] = []
        add_event(
            events,
            event_type="prenatal_consult",
            event_date=last_consultation_date,
            professional=professional,
            notes="Consulta mais recente importada da planilha-base.",
            metadata={"source": "registro_pac"},
        )

        raw = {
            "external_code": external_code,
            "name": name,
            "locality": locality,
            "risk_level": risk_level,
            "last_consultation_date": last_consultation_date,
            "gestational_age_label": gestational_age_label,
            "professional": professional,
            "active_search": active_search,
            "notes": notes,
        }
        records.append(
            {
                "source_system": "registro_pac",
                "source_key": make_source_key(patient),
                "row_hash": make_row_hash(raw),
                "patient": patient,
                "events": events,
                "raw": raw,
            }
        )
    return records


def parse_dados_sheet(workbook_bytes: bytes) -> list[dict]:
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    if "Dados" not in workbook.sheetnames:
        return []

    sheet = workbook["Dados"]
    top_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    bottom_headers = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))

    headers = []
    current_top = ""
    for top_value, bottom_value in zip(top_headers, bottom_headers):
        label, current_top = composite_header(top_value, bottom_value, current_top)
        headers.append(label)

    records = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        row_map = {slug_text(header): value for header, value in zip(headers, row) if clean_text(header)}
        name = clean_text(row_map.get("nome"))
        if not name or name == "#ERROR!":
            continue

        gestational_age_label = clean_text(row_map.get("estagio_gestacional")) or None
        gestational_weeks = extract_weeks(row_map.get("ig_em_semanas")) or extract_weeks(gestational_age_label)
        dum = parse_excel_date(row_map.get("dum"))
        dpp = parse_excel_date(row_map.get("dpp"))
        last_consultation_date = parse_excel_date(row_map.get("data_de_atendimento"))
        actual_birth_date = parse_excel_date(row_map.get("parto_e_nascimento_data")) or parse_excel_date(
            row_map.get("parto_e_nascimento_data_da_alta")
        )
        risk_level = normalize_risk(row_map.get("estratificacao_de_risco_registrar_o_resultado_da_estratificacao_e_os_tres_fatores_mais_importantes_identificados_resultado"))
        if risk_level == "Sem classificacao":
            risk_level = normalize_risk(row_map.get("vulnerabilidade"))

        patient = {
            "external_code": clean_text(row_map.get("numero_ordem")) or clean_text(row_map.get("prontuario")) or None,
            "cpf": digits_only(row_map.get("cpf_somente_numeros")) or None,
            "name": name,
            "birth_date": parse_excel_date(row_map.get("data_nascimento")),
            "mother_name": clean_text(row_map.get("nome_da_mae")) or None,
            "risk_level": risk_level,
            "gestational_age_label": gestational_age_label,
            "gestational_weeks": gestational_weeks,
            "dum": dum,
            "dpp": dpp,
            "actual_birth_date": actual_birth_date,
            "last_consultation_date": last_consultation_date,
            "last_professional": clean_text(row_map.get("responsavel_pelo_registro")) or None,
            "maternity_reference": clean_text(
                row_map.get("vinculacao_a_maternidade_nome_da_maternidade_atualizado_de_acordo_com_a_estratificacao_de_risco")
            )
            or clean_text(row_map.get("parto_e_nascimento_maternidade_nome"))
            or None,
            "high_risk_shared_care": parse_boolish(
                row_map.get("acompanhamento_no_pre_natal_de_alto_risco_compartilhamento_do_cuidado")
            ),
            "notes": clean_text(row_map.get("observacoes")) or None,
            "source": "pec_dados",
        }
        patient["status"] = infer_status(
            {
                "status": row_map.get("desfecho"),
                "desfecho": row_map.get("desfecho"),
                "actual_birth_date": actual_birth_date,
                "gestational_age_label": gestational_age_label,
            }
        )

        events: list[dict] = []
        consult_type = "puerperal_consult" if patient["status"] == "puerpera" else "prenatal_consult"
        add_event(
            events,
            event_type=consult_type,
            event_date=last_consultation_date,
            professional=patient.get("last_professional"),
            notes="Atendimento importado do PEC.",
            metadata={"source": "pec_dados"},
        )
        if clean_text(row_map.get("peso_em_kg")) or clean_text(row_map.get("estatura_em_m")):
            add_event(
                events,
                event_type="anthropometry",
                event_date=last_consultation_date,
                professional=patient.get("last_professional"),
                notes="Peso e/ou estatura registrados no PEC.",
                metadata={"source": "pec_dados"},
            )
        if clean_text(row_map.get("nivel_pressorico_pas_mmhg")) or clean_text(row_map.get("nivel_pressorico_pad_mmhg")):
            add_event(
                events,
                event_type="blood_pressure",
                event_date=last_consultation_date,
                professional=patient.get("last_professional"),
                notes="Pressao arterial registrada no PEC.",
                metadata={"source": "pec_dados"},
            )
        if parse_boolish(row_map.get("avaliacao_odontologica")):
            add_event(
                events,
                event_type="dental_visit",
                event_date=last_consultation_date,
                professional=patient.get("last_professional"),
                notes="Atividade odontologica importada do PEC.",
                metadata={"source": "pec_dados"},
            )
        tests_text = " ".join(
            clean_text(row_map.get(key))
            for key in (
                "testes_rapidos_previstos_para_o_trimestre",
                "exames_previstos_para_o_trimestre",
            )
            if row_map.get(key) is not None
        ).lower()
        if "realizado" in tests_text and patient["gestational_weeks"] and patient["gestational_weeks"] <= 13:
            add_event(
                events,
                event_type="first_trimester_tests",
                event_date=last_consultation_date,
                professional=patient.get("last_professional"),
                notes="Testes do primeiro trimestre importados do PEC.",
                metadata={"source": "pec_dados"},
            )
        if "realizado" in tests_text and patient["gestational_weeks"] and patient["gestational_weeks"] >= 28:
            add_event(
                events,
                event_type="third_trimester_tests",
                event_date=last_consultation_date,
                professional=patient.get("last_professional"),
                notes="Testes do terceiro trimestre importados do PEC.",
                metadata={"source": "pec_dados"},
            )
        add_event(
            events,
            event_type="delivery",
            event_date=actual_birth_date,
            professional=patient.get("last_professional"),
            notes="Parto importado do PEC.",
            metadata={"source": "pec_dados"},
        )

        raw = {
            "external_code": patient.get("external_code"),
            "cpf": patient.get("cpf"),
            "name": patient.get("name"),
            "birth_date": patient.get("birth_date"),
            "mother_name": patient.get("mother_name"),
            "risk_level": patient.get("risk_level"),
            "status": patient.get("status"),
            "gestational_age_label": patient.get("gestational_age_label"),
            "gestational_weeks": patient.get("gestational_weeks"),
            "dum": patient.get("dum"),
            "dpp": patient.get("dpp"),
            "actual_birth_date": patient.get("actual_birth_date"),
            "last_consultation_date": patient.get("last_consultation_date"),
            "last_professional": patient.get("last_professional"),
            "maternity_reference": patient.get("maternity_reference"),
            "high_risk_shared_care": patient.get("high_risk_shared_care"),
            "notes": patient.get("notes"),
        }
        records.append(
            {
                "source_system": "pec_dados",
                "source_key": make_source_key(patient),
                "row_hash": make_row_hash(raw),
                "patient": patient,
                "events": events,
                "raw": raw,
            }
        )
    return records


def parse_delimited_text(file_bytes: bytes) -> list[dict]:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    rows = list(csv.DictReader(StringIO(text), dialect=dialect))
    records = []
    for row in rows:
        normalized = {slug_text(key): value for key, value in row.items()}
        name = clean_text(normalized.get("nome") or normalized.get("name"))
        if not name:
            continue
        gestational_age_label = clean_text(normalized.get("idade_gestacional") or normalized.get("estagio_gestacional")) or None
        gestational_weeks = extract_weeks(normalized.get("ig_em_semanas")) or extract_weeks(gestational_age_label)
        patient = {
            "external_code": clean_text(normalized.get("numero") or normalized.get("prontuario")) or None,
            "cpf": digits_only(normalized.get("cpf")) or None,
            "name": name,
            "birth_date": parse_excel_date(normalized.get("data_nascimento")),
            "mother_name": clean_text(normalized.get("nome_da_mae")) or None,
            "locality": clean_text(normalized.get("localidade")) or None,
            "risk_level": normalize_risk(normalized.get("classificacao_risco") or normalized.get("risco")),
            "gestational_age_label": gestational_age_label,
            "gestational_weeks": gestational_weeks,
            "dum": parse_excel_date(normalized.get("dum")),
            "dpp": parse_excel_date(normalized.get("dpp")),
            "actual_birth_date": parse_excel_date(normalized.get("data_parto")),
            "last_consultation_date": parse_excel_date(normalized.get("ultima_consulta") or normalized.get("data_atendimento")),
            "last_professional": clean_text(normalized.get("profissional") or normalized.get("responsavel")) or None,
            "notes": clean_text(normalized.get("observacoes")) or None,
            "source": "arquivo_delimitado",
        }
        patient["status"] = infer_status(patient)
        raw = dict(patient)
        events: list[dict] = []
        add_event(
            events,
            event_type="puerperal_consult" if patient["status"] == "puerpera" else "prenatal_consult",
            event_date=patient.get("last_consultation_date"),
            professional=patient.get("last_professional"),
            notes="Atendimento importado do arquivo tabular.",
            metadata={"source": "arquivo_delimitado"},
        )
        add_event(
            events,
            event_type="delivery",
            event_date=patient.get("actual_birth_date"),
            professional=patient.get("last_professional"),
            notes="Parto importado do arquivo tabular.",
            metadata={"source": "arquivo_delimitado"},
        )
        records.append(
            {
                "source_system": "arquivo_delimitado",
                "source_key": make_source_key(patient),
                "row_hash": make_row_hash(raw),
                "patient": patient,
                "events": events,
                "raw": raw,
            }
        )
    return records


def parse_uploaded_records(filename: str, file_bytes: bytes, source_kind_hint: str | None = None) -> tuple[str, list[dict]]:
    hint = clean_text(source_kind_hint).lower()
    lower_name = filename.lower()
    if lower_name.endswith((".xlsx", ".xlsm")):
        if hint == "pec":
            return "pec_dados", parse_dados_sheet(file_bytes)
        if hint == "registro":
            return "registro_pac", parse_registro_sheet(file_bytes)
        dados_records = parse_dados_sheet(file_bytes)
        if dados_records:
            return "pec_dados", dados_records
        return "registro_pac", parse_registro_sheet(file_bytes)
    if lower_name.endswith((".csv", ".tsv", ".txt")):
        return "arquivo_delimitado", parse_delimited_text(file_bytes)
    raise ValueError("Formato de arquivo nao suportado. Use .xlsx, .xlsm, .csv, .tsv ou .txt.")


def safe_note_merge(current: str | None, incoming: str | None) -> str | None:
    current_text = clean_text(current)
    incoming_text = clean_text(incoming)
    if not incoming_text:
        return current_text or None
    if not current_text:
        return incoming_text
    if incoming_text in current_text:
        return current_text
    return f"{current_text} | {incoming_text}"


def resolve_patient_by_identity(connection, patient: dict) -> int | None:
    cpf = digits_only(patient.get("cpf"))
    if cpf:
        row = connection.execute("SELECT id FROM patients WHERE cpf = ? LIMIT 1", (cpf,)).fetchone()
        if row:
            return row["id"]

    cns = digits_only(patient.get("cns"))
    if cns:
        row = connection.execute("SELECT id FROM patients WHERE cns = ? LIMIT 1", (cns,)).fetchone()
        if row:
            return row["id"]

    external_code = clean_text(patient.get("external_code"))
    if external_code:
        row = connection.execute("SELECT id FROM patients WHERE external_code = ? LIMIT 1", (external_code,)).fetchone()
        if row:
            return row["id"]

    birth_date = patient.get("birth_date")
    if clean_text(patient.get("name")) and birth_date:
        row = connection.execute(
            "SELECT id FROM patients WHERE name = ? AND birth_date = ? LIMIT 1",
            (patient["name"], birth_date),
        ).fetchone()
        if row:
            return row["id"]

    mother_name = clean_text(patient.get("mother_name"))
    if clean_text(patient.get("name")) and mother_name:
        row = connection.execute(
            "SELECT id FROM patients WHERE name = ? AND mother_name = ? LIMIT 1",
            (patient["name"], mother_name),
        ).fetchone()
        if row:
            return row["id"]

    return None


def create_patient_from_import(connection, patient: dict, source_system: str) -> int:
    params = (
        patient.get("external_code"),
        patient.get("name"),
        patient.get("cpf"),
        patient.get("cns"),
        patient.get("birth_date"),
        patient.get("mother_name"),
        patient.get("locality"),
        patient.get("risk_level") or "Sem classificacao",
        patient.get("status") or "gestante",
        patient.get("gestational_weeks"),
        patient.get("gestational_age_label"),
        patient.get("dum"),
        patient.get("dpp"),
        patient.get("actual_birth_date"),
        patient.get("last_consultation_date"),
        patient.get("last_professional"),
        patient.get("maternity_reference"),
        patient.get("high_risk_shared_care", 0),
        patient.get("active_search", 0),
        source_system,
        patient.get("notes"),
    )
    if is_postgres_connection(connection):
        return connection.execute(
            """
            INSERT INTO patients (
                external_code, name, cpf, cns, birth_date, mother_name, locality,
                risk_level, status, gestational_weeks, gestational_age_label, dum, dpp,
                actual_birth_date, last_consultation_date, last_professional,
                maternity_reference, high_risk_shared_care, active_search, source, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            RETURNING id
            """,
            params,
        ).fetchone()["id"]
    return connection.execute(
        """
        INSERT INTO patients (
            external_code, name, cpf, cns, birth_date, mother_name, locality,
            risk_level, status, gestational_weeks, gestational_age_label, dum, dpp,
            actual_birth_date, last_consultation_date, last_professional,
            maternity_reference, high_risk_shared_care, active_search, source, notes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        params,
    ).lastrowid


def update_patient_from_import(connection, patient_id: int, patient: dict, source_system: str) -> bool:
    current = connection.execute("SELECT * FROM patients WHERE id = ?", (patient_id,)).fetchone()
    if not current:
        raise ValueError("Paciente nao encontrada durante importacao.")
    current_data = dict(current)

    merged = dict(current_data)
    for field in (
        "external_code",
        "cpf",
        "cns",
        "birth_date",
        "mother_name",
        "locality",
        "gestational_weeks",
        "gestational_age_label",
        "dum",
        "dpp",
        "actual_birth_date",
        "last_professional",
        "maternity_reference",
    ):
        if patient.get(field) not in (None, ""):
            merged[field] = patient.get(field)

    if patient.get("risk_level") and patient["risk_level"] != "Sem classificacao":
        merged["risk_level"] = patient["risk_level"]
    if patient.get("status"):
        merged["status"] = patient["status"]
    if patient.get("high_risk_shared_care"):
        merged["high_risk_shared_care"] = 1
    if patient.get("active_search"):
        merged["active_search"] = 1
    if patient.get("last_consultation_date"):
        current_last = current_data.get("last_consultation_date") or ""
        if not current_last or patient["last_consultation_date"] >= current_last:
            merged["last_consultation_date"] = patient["last_consultation_date"]
    merged["notes"] = safe_note_merge(current_data.get("notes"), patient.get("notes"))
    merged["source"] = source_system
    if merged.get("actual_birth_date"):
        merged["status"] = "puerpera"

    fields = [
        "external_code",
        "cpf",
        "cns",
        "birth_date",
        "mother_name",
        "locality",
        "risk_level",
        "status",
        "gestational_weeks",
        "gestational_age_label",
        "dum",
        "dpp",
        "actual_birth_date",
        "last_consultation_date",
        "last_professional",
        "maternity_reference",
        "high_risk_shared_care",
        "active_search",
        "source",
        "notes",
    ]
    changed = any(current_data.get(field) != merged.get(field) for field in fields)
    if not changed:
        return False

    values = [merged.get(field) for field in fields]
    values.append(patient_id)
    connection.execute(
        f"UPDATE patients SET {', '.join(f'{field} = ?' for field in fields)} WHERE id = ?",
        values,
    )
    return True


def upsert_source_link(connection, *, patient_id: int, source_system: str, source_key: str, row_hash: str, raw_payload: dict, import_run_id: int) -> None:
    existing = connection.execute(
        "SELECT id FROM patient_source_links WHERE source_system = ? AND source_key = ?",
        (source_system, source_key),
    ).fetchone()
    raw_json = json.dumps(raw_payload, ensure_ascii=False, sort_keys=True, default=str)
    if existing:
        connection.execute(
            """
            UPDATE patient_source_links
            SET patient_id = ?, source_hash = ?, raw_snapshot_json = ?, last_import_run_id = ?, last_seen_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (patient_id, row_hash, raw_json, import_run_id, existing["id"]),
        )
        return

    connection.execute(
        """
        INSERT INTO patient_source_links (
            patient_id, source_system, source_key, source_hash, raw_snapshot_json, last_import_run_id
        ) VALUES (?, ?, ?, ?, ?, ?)
        """,
        (patient_id, source_system, source_key, row_hash, raw_json, import_run_id),
    )


def sync_import_events(connection, *, patient_id: int, source_system: str, source_key: str, events: list[dict]) -> int:
    inserted = 0
    for event in events:
        event_date = event.get("event_date")
        if not event_date:
            continue
        event_key = build_event_key(source_system, source_key, event["event_type"], event_date)
        metadata = dict(event.get("metadata") or {})
        metadata["source_system"] = source_system
        metadata["source_key"] = source_key
        metadata_json = json.dumps(metadata, ensure_ascii=False, sort_keys=True)
        statement = (
            """
            INSERT INTO events (
                patient_id, event_type, event_date, professional, metadata_json, notes, source_event_key
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT (source_event_key) DO NOTHING
            """
            if is_postgres_connection(connection)
            else """
            INSERT OR IGNORE INTO events (
                patient_id, event_type, event_date, professional, metadata_json, notes, source_event_key
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """
        )
        cursor = connection.execute(
            statement,
            (
                patient_id,
                event["event_type"],
                event_date,
                event.get("professional"),
                metadata_json,
                event.get("notes"),
                event_key,
            ),
        )
        if cursor.rowcount:
            inserted += 1
    return inserted


def import_records(
    connection,
    *,
    filename: str,
    file_bytes: bytes,
    imported_by_user_id: int | None,
    source_kind_hint: str | None = None,
) -> dict:
    source_kind, records = parse_uploaded_records(filename, file_bytes, source_kind_hint=source_kind_hint)
    file_hash = hashlib.sha256(file_bytes).hexdigest()
    if is_postgres_connection(connection):
        run_id = connection.execute(
            """
            INSERT INTO import_runs (filename, source_kind, file_hash, imported_by_user_id, total_rows, status)
            VALUES (?, ?, ?, ?, ?, 'completed')
            RETURNING id
            """,
            (filename, source_kind, file_hash, imported_by_user_id, len(records)),
        ).fetchone()["id"]
    else:
        run_id = connection.execute(
            """
            INSERT INTO import_runs (filename, source_kind, file_hash, imported_by_user_id, total_rows, status)
            VALUES (?, ?, ?, ?, ?, 'completed')
            """,
            (filename, source_kind, file_hash, imported_by_user_id, len(records)),
        ).lastrowid

    inserted_patients = 0
    updated_patients = 0
    inserted_events = 0
    skipped_rows = 0

    for record in records:
        source_system = record["source_system"]
        source_key = record["source_key"]
        row_hash = record["row_hash"]
        patient = record["patient"]

        link = connection.execute(
            """
            SELECT patient_id, source_hash
            FROM patient_source_links
            WHERE source_system = ? AND source_key = ?
            """,
            (source_system, source_key),
        ).fetchone()

        patient_id: int | None = None
        if link:
            patient_id = link["patient_id"]
            if link["source_hash"] == row_hash:
                upsert_source_link(
                    connection,
                    patient_id=patient_id,
                    source_system=source_system,
                    source_key=source_key,
                    row_hash=row_hash,
                    raw_payload=record["raw"],
                    import_run_id=run_id,
                )
                skipped_rows += 1
                continue

        if patient_id is None:
            patient_id = resolve_patient_by_identity(connection, patient)

        if patient_id is None:
            patient_id = create_patient_from_import(connection, patient, source_system)
            inserted_patients += 1
        else:
            if update_patient_from_import(connection, patient_id, patient, source_system):
                updated_patients += 1

        upsert_source_link(
            connection,
            patient_id=patient_id,
            source_system=source_system,
            source_key=source_key,
            row_hash=row_hash,
            raw_payload=record["raw"],
            import_run_id=run_id,
        )
        inserted_events += sync_import_events(
            connection,
            patient_id=patient_id,
            source_system=source_system,
            source_key=source_key,
            events=record["events"],
        )

    summary = {
        "source_kind": source_kind,
        "total_rows": len(records),
        "inserted_patients": inserted_patients,
        "updated_patients": updated_patients,
        "inserted_events": inserted_events,
        "skipped_rows": skipped_rows,
    }
    connection.execute(
        """
        UPDATE import_runs
        SET inserted_patients = ?, updated_patients = ?, inserted_events = ?, skipped_rows = ?, summary_json = ?
        WHERE id = ?
        """,
        (
            inserted_patients,
            updated_patients,
            inserted_events,
            skipped_rows,
            json.dumps(summary, ensure_ascii=False),
            run_id,
        ),
    )
    connection.commit()
    return {"run_id": run_id, **summary}
